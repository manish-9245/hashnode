import streamlit as st
import json
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO

def split_key(key):
    if '.' not in key:
        return key, ""
    return key.rsplit('.', 1)

def parse_nested_json(value):
    if isinstance(value, str):
        try:
            return json.loads(value)
        except:
            return value
    return value

def extract_value(data, path):
    keys = path.split('.')
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return None
    return current

def process_json_list(json_list):
    """Convert list of single-key dictionaries to a dictionary"""
    result = {}
    for item in json_list:
        for key, value in item.items():
            result[key] = value
    return result

st.title("JSON Reconciliation Tool")

system1_file = st.file_uploader("Upload System 1 JSON", type=['json'])
system2_file = st.file_uploader("Upload System 2 JSON", type=['json'])
attr_paths_input = st.text_area("Enter attribute paths (one per line):")
attr_paths = [p.strip() for p in attr_paths_input.split('\n') if p.strip()]

if st.button("Generate Reconciliation Report") and system1_file and system2_file and attr_paths:
    try:
        # Load JSON files
        system1_list = json.load(system1_file)
        system2_list = json.load(system2_file)
        
        # Validate JSON format
        if not isinstance(system1_list, list):
            st.error("System 1 JSON must be a list of objects")
            st.stop()
        if not isinstance(system2_list, list):
            st.error("System 2 JSON must be a list of objects")
            st.stop()
        
        # Convert list format to dictionary
        system1_data = process_json_list(system1_list)
        system2_data = process_json_list(system2_list)
        
        # Process System 1
        sys1_processed = {}
        for key, value in system1_data.items():
            id_part, version_part = split_key(key)
            sys1_processed[key] = {
                'id': id_part,
                'version': version_part,
                'data': parse_nested_json(value)
            }
        
        # Process System 2
        sys2_processed = {}
        for key, value in system2_data.items():
            id_part, version_part = split_key(key)
            sys2_processed[key] = {
                'id': id_part,
                'version': version_part,
                'data': parse_nested_json(value)
            }
        
        # Identify matched and orphan keys
        all_keys = set(sys1_processed.keys()) | set(sys2_processed.keys())
        matched_keys = []
        orphans_sys1 = []
        orphans_sys2 = []
        
        for key in all_keys:
            if key in sys1_processed and key in sys2_processed:
                matched_keys.append(key)
            elif key in sys1_processed:
                orphans_sys1.append(key)
            else:
                orphans_sys2.append(key)
        
        # Create report structure with vertical stacking for matched keys
        report_rows = []
        
        # Process matched keys - show System1 then System2 vertically
        for key in matched_keys:
            id_val = sys1_processed[key]['id']
            ver_val = sys1_processed[key]['version']
            data1 = sys1_processed[key]['data']
            data2 = sys2_processed[key]['data']
            
            # Create row for System1
            row_sys1 = {
                'system': 'System1',
                'id': id_val,
                'version': ver_val,
                'is_orphan': False
            }
            
            # Create row for System2
            row_sys2 = {
                'system': 'System2',
                'id': id_val,
                'version': ver_val,
                'is_orphan': False
            }
            
            # Add attribute values
            for path in attr_paths:
                val1 = extract_value(data1, path)
                val2 = extract_value(data2, path)
                
                # Determine if values match
                if val1 == val2:
                    row_sys1[path] = {'value': val1, 'color': 'green'}
                    row_sys2[path] = {'value': val2, 'color': 'green'}
                else:
                    row_sys1[path] = {'value': val1, 'color': 'red'}
                    row_sys2[path] = {'value': val2, 'color': 'red'}
            
            report_rows.append(row_sys1)
            report_rows.append(row_sys2)
        
        # Process orphans - show at bottom
        for key in orphans_sys1:
            id_val = sys1_processed[key]['id']
            ver_val = sys1_processed[key]['version']
            data1 = sys1_processed[key]['data']
            
            row = {
                'system': 'System1',
                'id': id_val,
                'version': ver_val,
                'is_orphan': True
            }
            
            for path in attr_paths:
                val = extract_value(data1, path)
                row[path] = {'value': val, 'color': 'yellow'}
            
            report_rows.append(row)
        
        for key in orphans_sys2:
            id_val = sys2_processed[key]['id']
            ver_val = sys2_processed[key]['version']
            data2 = sys2_processed[key]['data']
            
            row = {
                'system': 'System2',
                'id': id_val,
                'version': ver_val,
                'is_orphan': True
            }
            
            for path in attr_paths:
                val = extract_value(data2, path)
                row[path] = {'value': val, 'color': 'yellow'}
            
            report_rows.append(row)
        
        # Generate Excel Report
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create headers
        headers = ['System', 'ID', 'Version'] + attr_paths
        ws.append(headers)
        
        # Define color fills
        color_fills = {
            'green': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        }
        
        # Add data rows
        for row in report_rows:
            data_row = [
                row['system'],
                row['id'],
                row['version']
            ]
            
            for path in attr_paths:
                data_row.append(str(row[path]['value']) if row[path]['value'] is not None else '')
            
            ws.append(data_row)
            
            # Apply coloring
            row_idx = ws.max_row
            
            # Color entire row for orphans
            if row['is_orphan']:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = color_fills['yellow']
            else:
                # Color attribute cells for matched rows
                for col_idx, path in enumerate(attr_paths, start=4):  # Start at column 4
                    color = row[path]['color']
                    ws.cell(row=row_idx, column=col_idx).fill = color_fills[color]
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    value_length = len(str(cell.value)) if cell.value else 0
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
        
        # Set wrap text for long values
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Create Excel file in memory
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate HTML Report
        html_content = """
        <html>
        <head>
            <title>Reconciliation Report</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 20px;
                }
                h1 {
                    color: #2c3e50;
                    text-align: center;
                }
                table {
                    border-collapse: collapse;
                    width: 100%;
                    margin-bottom: 20px;
                    font-size: 12px;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                    vertical-align: top;
                }
                th {
                    background-color: #f2f2f2;
                    position: sticky;
                    top: 0;
                    z-index: 10;
                }
                tr.yellow td {
                    background-color: #FFFF00 !important;
                }
                .value-cell {
                    max-width: 300px;
                    overflow: auto;
                }
                .system-row {
                    border-bottom: 2px solid #333;
                }
            </style>
        </head>
        <body>
            <h1>Reconciliation Report</h1>
            <table>
                <thead>
                    <tr>
                        <th>System</th>
                        <th>ID</th>
                        <th>Version</th>
        """
        
        # Add attribute headers
        for path in attr_paths:
            html_content += f"<th>{path}</th>"
        
        html_content += """
                    </tr>
                </thead>
                <tbody>
        """
        
        # Add table rows
        for row in report_rows:
            row_class = "yellow" if row['is_orphan'] else ""
            html_content += f"<tr class='{row_class}'>"
            html_content += f"<td>{row['system']}</td>"
            html_content += f"<td>{row['id']}</td>"
            html_content += f"<td>{row['version']}</td>"
            
            for path in attr_paths:
                value = row[path]['value']
                color = row[path]['color'] if not row['is_orphan'] else 'yellow'
                
                if row['is_orphan']:
                    html_content += f"<td class='value-cell'>{str(value) if value is not None else ''}</td>"
                else:
                    html_content += f"<td class='value-cell' style='background-color: {color};'>{str(value) if value is not None else ''}</td>"
            
            html_content += "</tr>"
            
            # Add separator after System2 rows for matched keys
            if not row['is_orphan'] and row['system'] == 'System2':
                html_content += "<tr class='system-row'><td colspan='100%'></td></tr>"
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        # Create HTML file in memory
        html_buffer = BytesIO()
        html_buffer.write(html_content.encode('utf-8'))
        html_buffer.seek(0)
        
        # Provide download links
        st.success("Report generated successfully!")
        
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="Download Excel Report",
                data=excel_buffer,
                file_name="reconciliation_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="Download HTML Report",
                data=html_buffer,
                file_name="reconciliation_report.html",
                mime="text/html"
            )
        
        # Show preview of first 5 keys
        st.subheader("Preview (First 5 Keys)")
        preview_data = []
        for i, row in enumerate(report_rows):
            if i >= 10:  # Show up to 10 rows (5 keys)
                break
            
            preview_row = {
                "System": row['system'],
                "ID": row['id'],
                "Version": row['version']
            }
            
            for path in attr_paths:
                value = row[path]['value']
                preview_row[path] = str(value)[:50] + "..." if value and len(str(value)) > 50 else str(value)
            
            preview_data.append(preview_row)
        
        st.table(preview_data)
    
    except json.JSONDecodeError:
        st.error("Invalid JSON file format. Please upload valid JSON files.")
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
